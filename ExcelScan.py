import openpyxl
import pywikibot

def getSearchCount(site, name, maxSearchNumber):
    if maxSearchNumber < 200:
        maxSearchNumber = 500
    else:
        maxSearchNumber = 2*maxSearchNumber
    results = list(site.search("\""+name+"\"", total=maxSearchNumber, namespaces=[0]))
    #print(f"Treffer für '{name}' im Artikelnamensraum: {len(results)}")
    return len(results)

# Lade Excel-Datei
dateiname = "../Mineralnamen.xlsx"  # Passe den Dateinamen an
wb = openpyxl.load_workbook(dateiname)
ws = wb.active

# Richte Pywikibot-Site-Objekte ein
site_de = pywikibot.Site("de", "wikipedia")
repo = site_de.data_repository()

ws.cell(row=1, column=3).value = "Links"
ws.cell(row=1, column=4).value = "Sprachen"
ws.cell(row=1, column=5).value = "Suchtreffer"

# Durchlaufe alle Zeilen ab der zweiten (Kopfzeile überspringen)
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    name = row[0].value         # Spalte A: Name
    wikidata_id = row[1].value  # Spalte B: Wikidata-ID
    link_count = 0
    language_count = 0
    search_count = 0
    
    # Anzahl interner Links in deutscher Wikipedia
    try:
        page = pywikibot.Page(site_de, name)
        backlinks = list(page.backlinks(namespaces=[0]))  # Nur Hauptnamensraum
        link_count = len(backlinks)
    except Exception as e:
        print(f"Fehler bei Seite '{name}': {e}")
        link_count = None

    search_count = getSearchCount(site_de, name, link_count)

    # Anzahl Sprachversionen über Wikidata
    if isinstance(wikidata_id, str) and wikidata_id.startswith("Q"):
        try:
            item = pywikibot.ItemPage(repo, wikidata_id)
            item.get()  # Daten abrufen
            wikipedia_links = {
                key: link
                for key, link in item.sitelinks.items()
                if key.endswith('wiki') and not key.startswith(('commons', 'wikidata', 'species', 'meta'))
            }
            language_count = len(wikipedia_links)

        except Exception as e:
            print(f"Fehler bei Wikidata-ID '{wikidata_id}': {e}")
            language_count = None
    else:
        language_count = None
        
    print(f"{i}. name={name}, wikidata_id={wikidata_id}, links={link_count}, languages={language_count}, search_count={search_count}")
    # Werte in Spalte C (2 = Index 2) und D (3 = Index 3) schreiben
    ws.cell(row=i, column=3).value = link_count
    ws.cell(row=i, column=4).value = language_count
    ws.cell(row=i, column=5).value = search_count
    
    if i%100 == 0:
        print("save")
        wb.save("../Mineralnamen_aktualisiert.xlsx")
    #    break

# Speichere Datei
wb.save("../Mineralnamen_aktualisiert.xlsx")
print("Fertig! Datei gespeichert als 'Mineralnamen_aktualisiert.xlsx'")
